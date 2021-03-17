import contextlib
import openpyxl
from os import path


@contextlib.contextmanager
def context_manager_for_exel(filename):
    if path.exists(filename):
        workbook = openpyxl.load_workbook(filename)
    else:
        workbook = openpyxl.Workbook()
        workbook.save(filename)
    try:
        yield workbook
        workbook.save(filename)
    except BaseException as exc:
        print(exc.args)


class ContextManagerForExcel:

    def __init__(self, filename):
        self.filename = filename
        if path.exists(filename):
            self.workbook = openpyxl.load_workbook(filename)
        else:
            self.workbook = openpyxl.Workbook()
            self.workbook.save(filename)

    def __enter__(self):
        return self.workbook

    def __exit__(self, exc_type, exc_val, exc_tb):
        if exc_type:
            print(exc_type, exc_val, exc_tb)
        else:
            self.workbook.save(self.filename)
        return True
